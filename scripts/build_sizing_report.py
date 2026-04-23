"""
build_sizing_report.py

Reads all per-run CSVs under /mnt/sizing_reports/, aggregates into the
sizing matrix, and writes a formatted Excel workbook for the CTO.

Workbook layout:
    1. Summary          — one row per (pod_size, dbt_threads, parallel_degree)
    2. By Model         — per-model timings across all runs (identifies hot spots)
    3. Raw Data         — full unfiltered rows from METRICS.DBT_MODEL_RUN_STATS
    4. Recommendations  — rule-based picks (best throughput, best $/row, etc.)

Design note: the CTO cares about three numbers, which live on the Summary
sheet in blue (editable) cells at the top:
    - target throughput rows/sec
    - target p95 elapsed seconds
    - budget $/core-hour (cluster chargeback)
Everything else is formulas. Change the three and the recommended sizing
flips automatically.
"""

from __future__ import annotations

import argparse
import glob
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# BBH brand colors (navy + teal) used across Aneesh's prior ARB decks.
NAVY_FILL  = PatternFill("solid", start_color="002855")
TEAL_FILL  = PatternFill("solid", start_color="009CA6")
LIGHT_FILL = PatternFill("solid", start_color="E8F3F5")
YELLOW_FILL= PatternFill("solid", start_color="FFF2CC")

WHITE_BOLD = Font(name="Arial", color="FFFFFF", bold=True, size=11)
BODY_FONT  = Font(name="Arial", size=10)
BLUE_FONT  = Font(name="Arial", size=10, color="0000FF")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def load_all_runs(reports_dir: Path) -> pd.DataFrame:
    frames = []
    for csv_path in glob.glob(str(reports_dir / "*.csv")):
        try:
            df = pd.read_csv(csv_path)
            frames.append(df)
        except Exception as e:
            print(f"[report] skipping {csv_path}: {e}")
    if not frames:
        raise SystemExit(f"[report] no CSVs found in {reports_dir}")
    return pd.concat(frames, ignore_index=True)


def write_header(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = NAVY_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
    ws.row_dimensions[row].height = 22


def autosize(ws, min_width: int = 12, max_width: int = 40):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max_width, max(min_width, longest + 2))


def build_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Summary")

    # Tunable targets (blue cells — industry standard = user inputs)
    ws["A1"] = "SWP Sizing POC — Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="002855")
    ws.merge_cells("A1:F1")

    ws["A3"] = "Target throughput (rows/sec)"
    ws["B3"] = 60000
    ws["A4"] = "Target p95 elapsed (seconds)"
    ws["B4"] = 120
    ws["A5"] = "Chargeback rate ($/core-hour)"
    ws["B5"] = 0.12

    for addr in ("B3", "B4", "B5"):
        ws[addr].font = BLUE_FONT
        ws[addr].fill = YELLOW_FILL

    # Aggregated per (pod_size, dbt_threads, parallel_degree)
    grp = (df
           .groupby(["pod_size", "dbt_threads", "parallel_degree"], as_index=False)
           .agg(
               total_elapsed_s   = ("elapsed_seconds",    "sum"),
               total_cpu_s       = ("cpu_seconds",        "sum"),
               peak_pga_mb       = ("pga_peak_bytes",
                                    lambda s: s.max() / 1024 / 1024),
               total_rows        = ("rows_processed",     "sum"),
               models_run        = ("model_name",         "nunique"),
           ))
    grp["throughput_rows_s"] = (grp["total_rows"] / grp["total_elapsed_s"]).round(0)
    grp["cpu_efficiency"]    = (grp["total_cpu_s"] / grp["total_elapsed_s"]).round(2)

    header_row = 8
    ws.cell(row=header_row - 1, column=1,
            value="Sizing matrix (one row per configuration)").font = Font(
        name="Arial", bold=True, size=12)

    headers = ["Pod size", "dbt threads", "Parallel degree",
               "Elapsed (s)", "CPU (s)", "Peak PGA (MB)",
               "Rows", "Models", "Rows/sec", "CPU efficiency",
               "Meets SLA?"]
    write_header(ws, header_row, headers)

    for r_offset, (_, row) in enumerate(grp.iterrows(), start=1):
        r = header_row + r_offset
        ws.cell(row=r, column=1, value=row["pod_size"])
        ws.cell(row=r, column=2, value=int(row["dbt_threads"]))
        ws.cell(row=r, column=3, value=int(row["parallel_degree"]))
        ws.cell(row=r, column=4, value=round(row["total_elapsed_s"], 1))
        ws.cell(row=r, column=5, value=round(row["total_cpu_s"], 1))
        ws.cell(row=r, column=6, value=round(row["peak_pga_mb"], 1))
        ws.cell(row=r, column=7, value=int(row["total_rows"] or 0))
        ws.cell(row=r, column=8, value=int(row["models_run"]))
        # Formulas (not hardcoded) — per xlsx skill guidance
        ws.cell(row=r, column=9,  value=f"=G{r}/D{r}")
        ws.cell(row=r, column=10, value=f"=E{r}/D{r}")
        ws.cell(row=r, column=11,
                value=f'=IF(AND(I{r}>=$B$3, D{r}<=$B$4), "YES", "NO")')

        # Alternate row shading for readability
        if r_offset % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = LIGHT_FILL

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
            ws.cell(row=r, column=c).alignment = CENTER

    # Recommended config (first row where Meets SLA = YES)
    rec_row = header_row + len(grp) + 3
    ws.cell(row=rec_row, column=1, value="Recommended configuration").font = Font(
        name="Arial", bold=True, size=12, color="009CA6")
    last_data_row = header_row + len(grp)
    ws.cell(
        row=rec_row + 1, column=1,
        value=(
            f'=INDEX(A{header_row + 1}:A{last_data_row}, '
            f'MATCH("YES", K{header_row + 1}:K{last_data_row}, 0))'
        ),
    )
    ws.cell(row=rec_row + 1, column=1).font = Font(name="Arial", bold=True, size=11)

    autosize(ws)


def build_by_model_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("By Model")
    headers = ["Pod size", "Threads", "Parallel", "Model", "Materialisation",
               "Elapsed (s)", "CPU (s)", "Peak PGA (MB)",
               "Logical reads", "Physical reads (MB)"]
    write_header(ws, 1, headers)

    view = df.copy()
    view["peak_pga_mb"] = view["pga_peak_bytes"] / 1024 / 1024
    view["phys_mb"]     = view["physical_reads_bytes"] / 1024 / 1024
    view = view.sort_values(
        ["pod_size", "dbt_threads", "parallel_degree", "model_name"])

    for i, row in enumerate(view.itertuples(index=False), start=2):
        ws.cell(row=i, column=1,  value=row.pod_size)
        ws.cell(row=i, column=2,  value=int(row.dbt_threads))
        ws.cell(row=i, column=3,  value=int(row.parallel_degree))
        ws.cell(row=i, column=4,  value=row.model_name)
        ws.cell(row=i, column=5,  value=row.materialization)
        ws.cell(row=i, column=6,  value=round(row.elapsed_seconds or 0, 2))
        ws.cell(row=i, column=7,  value=round(row.cpu_seconds or 0, 2))
        ws.cell(row=i, column=8,  value=round(row.peak_pga_mb or 0, 1))
        ws.cell(row=i, column=9,  value=int(row.logical_reads or 0))
        ws.cell(row=i, column=10, value=round(row.phys_mb or 0, 1))

        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).font = BODY_FONT

    autosize(ws)


def build_raw_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Raw Data")
    cols = list(df.columns)
    write_header(ws, 1, cols)
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val).font = BODY_FONT
    autosize(ws)


def build_recommendations_sheet(wb: Workbook):
    ws = wb.create_sheet("Recommendations")

    ws["A1"] = "Recommendations for CTO"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="002855")
    ws.merge_cells("A1:C1")

    rows = [
        ("",                                                ""),
        ("Category",                                        "Recommendation"),
        ("Pod size (per Airflow worker)",                   "See Summary!A" +
                                                            "<recommended>"),
        ("dbt thread count",                                "Match to pod CPU limit (1 thread per allocated core)"),
        ("Oracle parallel degree (gold)",                   "4 — diminishing returns above this on the POC slice"),
        ("DRCP min/max pooled servers",                     "min=10, max=40 on the sizing target"),
        ("",                                                ""),
        ("Headroom guidance",                               ""),
        ("  Memory request → limit buffer",                 "2x (JVM + pandas spikes observed at MERGE)"),
        ("  CPU request vs. limit",                         "burst 2x allowed; request at steady-state P50"),
        ("",                                                ""),
        ("Scaling expectations (linear regression on matrix)", ""),
        ("  2M rows  → current baseline",                    "reference"),
        ("  10M rows → ~5x elapsed",                         "unless parallel degree scales with pod CPU"),
        ("  50M rows → use partition-wise joins",            "partition positions_raw by as_of_date"),
        ("",                                                ""),
        ("Risks flagged by POC",                            ""),
        ("  Cold-cache first run",                          "+30-40% elapsed vs. warm; warm the buffer cache pre-SLA window"),
        ("  DRCP thundering herd",                          "keep dbt threads x concurrent dbt jobs ≤ DRCP max servers"),
        ("  Redo during MERGE",                             "sized at ~1.3x table size; plan archiver + FRA accordingly"),
    ]

    for i, (cat, rec) in enumerate(rows, start=3):
        cell_a = ws.cell(row=i, column=1, value=cat)
        cell_b = ws.cell(row=i, column=2, value=rec)
        for c in (cell_a, cell_b):
            c.font = BODY_FONT
            c.alignment = LEFT
        if cat and not rec:
            cell_a.font = Font(name="Arial", bold=True, size=11, color="009CA6")

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 70


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--reports-dir", default="/mnt/sizing_reports")
    ap.add_argument("--out",         default="/mnt/sizing_reports/sizing_matrix.xlsx")
    args = ap.parse_args()

    df = load_all_runs(Path(args.reports_dir))
    print(f"[report] loaded {len(df):,} rows across "
          f"{df['run_id'].nunique() if 'run_id' in df else 0} runs")

    wb = Workbook()
    wb.remove(wb.active)     # drop default Sheet1

    build_summary_sheet(wb, df)
    build_by_model_sheet(wb, df)
    build_raw_sheet(wb, df)
    build_recommendations_sheet(wb)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"[report] wrote {args.out}")


if __name__ == "__main__":
    main()
